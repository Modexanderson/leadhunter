"""
Export Manager - Save leads to CSV, Excel, JSON
"""

import json
import csv
import logging
from pathlib import Path
from datetime import datetime
from dataclasses import asdict

logger = logging.getLogger(__name__)


class ExportManager:

    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)

    def _timestamp(self) -> str:
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def to_dict_list(self, leads: list) -> list[dict]:
        result = []
        for lead in leads:
            d = asdict(lead)
            d["social_links"] = json.dumps(d["social_links"])
            d["tech_stack"] = ", ".join(d["tech_stack"])
            result.append(d)
        return result

    def export_csv(self, leads: list, filename: str = None) -> Path:
        filename = filename or f"leads_{self._timestamp()}.csv"
        path = self.export_dir / filename

        data = self.to_dict_list(leads)
        if not data:
            return path

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

        logger.info(f"Exported {len(leads)} leads to {path}")
        return path

    def export_excel(self, leads: list, filename: str = None) -> Path:
        filename = filename or f"leads_{self._timestamp()}.xlsx"
        path = self.export_dir / filename

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            data = self.to_dict_list(leads)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads"

            if not data:
                wb.save(path)
                return path

            headers = list(data[0].keys())

            # Header styling
            header_fill = PatternFill("solid", fgColor="1a1a2e")
            header_font = Font(color="FFFFFF", bold=True)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Score color coding
            score_col = headers.index("score") + 1 if "score" in headers else None

            for row_idx, row in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row[key])

                    # Color rows by score
                    if score_col:
                        score = row.get("score", 0)
                        if score >= 75:
                            cell.fill = PatternFill("solid", fgColor="d4edda")  # Green
                        elif score >= 50:
                            cell.fill = PatternFill("solid", fgColor="fff3cd")  # Yellow
                        elif score < 30:
                            cell.fill = PatternFill("solid", fgColor="f8d7da")  # Red

            # Auto-width columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            wb.save(path)
            logger.info(f"Exported {len(leads)} leads to Excel: {path}")

        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self.export_csv(leads, filename.replace(".xlsx", ".csv"))

        return path

    def export_json(self, leads: list, filename: str = None) -> Path:
        filename = filename or f"leads_{self._timestamp()}.json"
        path = self.export_dir / filename
        data = self.to_dict_list(leads)
        path.write_text(json.dumps(data, indent=2))
        return path

    def load_json(self, path: str) -> list:
        """Load previously saved leads"""
        from core.scraper import Lead
        data = json.loads(Path(path).read_text())
        leads = []
        for d in data:
            if isinstance(d.get("tech_stack"), str):
                d["tech_stack"] = [x.strip() for x in d["tech_stack"].split(",") if x.strip()]
            if isinstance(d.get("social_links"), str):
                try:
                    d["social_links"] = json.loads(d["social_links"])
                except:
                    d["social_links"] = {}
            leads.append(Lead(**d))
        return leads
